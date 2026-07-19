import os
import json
import sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def summarize_to_academic_table_v2():
    PROJECT_ROOT = Path(__file__).resolve().parents[2]
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    from project_config import EXTRACTOR_OUTPUT_ROOT

    base_dir = EXTRACTOR_OUTPUT_ROOT
    all_rows = []

    if not os.path.exists(base_dir):
        print(f"错误：找不到指定的目录 -> {base_dir}")
        return

    print("1. 开始扫描并解析所有案件的 JSON 文件...")
    for case_folder in os.listdir(base_dir):
        case_path = os.path.join(base_dir, case_folder)

        if os.path.isdir(case_path):
            json_file_path = os.path.join(case_path, "evaluation", "evaluation_summary_cross_agent.json")

            if os.path.exists(json_file_path):
                try:
                    with open(json_file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    event_name = data.get("event_name", case_folder)

                    # 提取全局平均
                    global_avg = data.get("global_averages", {})
                    if global_avg:
                        all_rows.append({
                            "Case": event_name, "Model": "Global Average",
                            "Accuracy": global_avg.get("accuracy_avg"),
                            "Logic": global_avg.get("logic_avg"),
                            "Professionalism": global_avg.get("professionalism_avg"),
                            "Actionability": global_avg.get("actionability_avg"),
                            "Overall": global_avg.get("overall_avg")
                        })

                    # 提取各模型得分
                    detailed_scores = data.get("detailed_model_scores", {})
                    for model_name, scores in detailed_scores.items():
                        all_rows.append({
                            "Case": event_name, "Model": model_name,
                            "Accuracy": scores.get("accuracy_score"),
                            "Logic": scores.get("logic_score"),
                            "Professionalism": scores.get("professionalism_score"),
                            "Actionability": scores.get("actionability_score"),
                            "Overall": scores.get("overall_score")
                        })
                except Exception as e:
                    print(f"解析文件失败 [{json_file_path}]: {e}")

    if not all_rows:
        print("未提取到任何有效数据。")
        return

    # 2. 转换为 DataFrame
    raw_df = pd.DataFrame(all_rows)

    # 3. 数据透视：将 Case 转换到列上，Model 留在行上
    print("2. 正在进行学术论文格式转换与指标对齐...")
    academic_df = raw_df.pivot(index="Model", columns="Case",
                               values=["Accuracy", "Logic", "Professionalism", "Actionability", "Overall"])

    # 调换列的层级，让“案件名”在第一层，“指标名”在第二层
    academic_df = academic_df.swaplevel(0, 1, axis=1)

    # 【核心修改点】显式定义指标顺序，确保 Overall 永远在每个事件的最后一列
    metric_order = ["Accuracy", "Logic", "Professionalism", "Actionability", "Overall"]

    # 获取当前所有的案件名称并排序
    cases_order = sorted(list(raw_df["Case"].unique()))

    # 构建多级列索引的完美顺序：[(Case1, Accuracy), (Case1, Logic) ... (Case2, Overall)]
    custom_columns_order = [(c, m) for c in cases_order for m in metric_order]
    academic_df = academic_df.reindex(columns=custom_columns_order)

    # 让 "Global Average"（全局平均行）显示在表格的最底部
    if "Global Average" in academic_df.index:
        models_order = [idx for idx in academic_df.index if idx != "Global Average"] + ["Global Average"]
        academic_df = academic_df.reindex(models_order)

    # 4. 写入 Excel 并进行样式美化
    output_excel = os.path.join(base_dir, "案件得分汇总表.xlsx")
    print(f"3. 正在写入并美化 Excel 文件...")

    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        # Sheet 1: 原始明细数据
        raw_df.to_excel(writer, sheet_name="原始得分明细", index=False)

        # Sheet 2: 论文最终表格
        academic_df.to_excel(writer, sheet_name="论文最终表格")

        # 获取 openpyxl 工作表对象进行样式精修
        workbook = writer.book
        worksheet = writer.sheets["论文最终表格"]

        # ---- 样式定义 ----
        # 字体
        font_base = Font(name="Times New Roman", size=10.5)
        font_header_1 = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")  # 一级表头白色
        font_header_2 = Font(name="Times New Roman", size=10, bold=True)  # 二级表头
        font_bold_row = Font(name="Times New Roman", size=10.5, bold=True)  # 全局平均加粗

        # 填充颜色
        fill_header_1 = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # 稳重学术蓝
        fill_overall_col = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")  # Overall列浅色高亮
        fill_avg_row = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")  # 全局平均行高亮

        # 对齐
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")

        # 框线 (学术准三线表风格)
        thin_side = Side(border_style="thin", color="D3D3D3")
        thick_top_side = Side(border_style="medium", color="000000")
        double_bottom_side = Side(border_style="double", color="000000")

        border_header_1 = Border(top=thick_top_side, bottom=thin_side, left=thin_side, right=thin_side)
        border_header_2 = Border(bottom=thick_top_side, left=thin_side, right=thin_side)
        border_data = Border(bottom=thin_side, left=thin_side, right=thin_side)
        border_total_row = Border(top=thin_side, bottom=double_bottom_side, left=thin_side, right=thin_side)

        # ---- 应用样式 ----
        # 1. 格式化表头行
        # 第一行 (案件名称)
        for cell in worksheet[1]:
            cell.font = font_header_1
            cell.fill = fill_header_1
            cell.alignment = align_center
            cell.border = border_header_1

        # 第二行 (指标名称)
        for cell in worksheet[2]:
            cell.font = font_header_2
            cell.alignment = align_center
            cell.border = border_header_2

        # 2. 格式化数据行
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        for r in range(3, max_row + 1):
            is_avg_row = (worksheet.cell(row=r, column=1).value == "Global Average")

            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                cell.font = font_base

                # 行首（模型名称）左对齐，其余数据居中
                if c == 1:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_center
                    # 尝试保留两位小数
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00' if is_avg_row else '0.0'

                # 背景色及边框逻辑
                if is_avg_row:
                    cell.font = font_bold_row
                    cell.fill = fill_avg_row
                    cell.border = border_total_row
                else:
                    cell.border = border_data
                    # 如果是 Overall 列，单独加一个浅色背景便于论文中辨识
                    if c > 1 and worksheet.cell(row=2, column=c).value == "Overall":
                        cell.fill = fill_overall_col

        # 3. 自动调整列宽
        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 22
        worksheet.column_dimensions['A'].width = 24  # 模型列加宽

        for col in range(2, max_col + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 12

    print(f"\n【学术汇总表完美生成！】")
    print(f"结果文件已保存至: {output_excel}")
    print(f"现在每个事件下的最后一列已强制对齐为 [Overall] 综合得分。")


if __name__ == "__main__":
    summarize_to_academic_table_v2()
